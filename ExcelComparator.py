import tkinter as tk
from customtkinter import *
from tkinter import filedialog, messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import os
import sys
import subprocess


class ExcelDiffTool:
    #colors
    white1 = "#FFFFFF"
    white2 = "#e9e9e3"
    white3 = "#d4d4ca"
    white4 = "#b4b4aa"
    white5 = "#999990"
    white6 = "#7A7A75"
    black1 = "#000000"
    black2 = "#1a1f18"
    black3 = "#252c22"
    black4 = "#353F31"
    green1 = "#10ca86"
    green2 = "#0ead6b"
    green3 = "#0c974d"
    green4 = "#0B7945"
    green5 = "#175808"
    green6 = "#064213"
    blue1 = "#bffff1"
    blue2 = "#74f2ce"
    blue3 = "#63ceaf"

    def __init__(self, master):
        set_appearance_mode("light")
        set_default_color_theme("green")

        self.master = master
        self.master.title("Excel Diff Tool")#title of the app
        self.master.geometry("465x320")
        self.master.configure(bg = self.white2)
        self.master.minsize(465, 320)   

        self.files = []

        # hint
        self.label_drag_hint = CTkLabel(master, text="Drag and drop Excel files (.xlsx) below \n Or use 'Add Files' button", text_color=self.white5, 
                                        font=CTkFont(size=15, weight="bold"))
        self.label_drag_hint.pack(pady=(10, 5))

        # frame with list
        self.list_frame = CTkFrame(master, corner_radius=10, fg_color=self.white3)
        self.list_frame.pack(fill="y", expand=True, padx=30, pady=(0, 0))

        # use textbox instead of listbox (more flexible)
        self.textbox = CTkTextbox(self.list_frame, wrap="none", fg_color=self.white1, width=450, height=100)
        self.textbox.pack(fill="both", expand=True, padx=5, pady=5)
        self.textbox.configure(state="disabled")
        # dnd
        self.textbox.drop_target_register(DND_FILES)
        self.textbox.dnd_bind('<<Drop>>', self.drop)

        # buttons
        btn_frame = CTkFrame(master, fg_color=self.white3)
        btn_frame.pack(padx=50, pady=(10, 0))

        self.add_btn = CTkButton(btn_frame, text="Add Files", command=self.add_files, height = 36, width = 110, fg_color=self.green1, hover_color=self.green2, 
                                 font=CTkFont(size=15, weight="bold"), text_color=self.green6)
        self.add_btn.pack(anchor="center", side ="left", padx=(10, 6), pady=6)

        self.remove_btn = CTkButton(btn_frame, text="Remove Last", command=self.remove_last, height = 36, width = 110, fg_color=self.green1, hover_color=self.green2,
                                    font=CTkFont(size=15, weight="bold"), text_color=self.green6)
        self.remove_btn.pack(anchor="center", side ="left", padx=(0, 6), pady=6)

        self.clear_btn = CTkButton(btn_frame, text="Clear All", command=self.clear_files, height = 36, width = 110, fg_color=self.green1, hover_color=self.green2,
                                   font=CTkFont(size=15, weight="bold"), text_color=self.green6)
        self.clear_btn.pack(anchor="center", side ="left", padx=(0, 10), pady=6)

        self.compare_btn = CTkButton(master, text="Compare", fg_color=self.blue2, hover_color=self.blue3, command=self.compare, width=150, height=50,
                                     font=CTkFont(size=20, weight="bold"), text_color=self.green6)
        self.compare_btn.pack(pady=(20, 20))

    # TODO: MAKE EXCEL SHEET PRETTIER WITH BETTER FORMATTING, COMMENT CELLS
    # TODO: ADD PROPER NAME AND LOGO
    
    def update_textbox(self):
        self.textbox.configure(state="normal")
        self.textbox.delete("1.0", "end")
        if self.files:
            for f in self.files:
                self.textbox.insert("end", f"{f}\n")
        else:
            self.textbox.insert("end", "")
        self.textbox.configure(state="disabled")

    def drop(self, event):
        paths = self.master.tk.splitlist(event.data)
        added = False
        for p in paths:
            if p.endswith('.xlsx') and p not in self.files:
                self.files.append(p)
                added = True
        if added:
            self.update_textbox()
        return "break"

    def add_files(self):
        paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
        for p in paths:
            if p.endswith('.xlsx') and p not in self.files:
                self.files.append(p)
        self.update_textbox()

    def remove_last(self):
        if self.files:
            self.files.pop()
            self.update_textbox()

    def clear_files(self):
        if messagebox.askyesno("Confirm", "Clear all files?"):
            self.files.clear()
            self.update_textbox()

    def compare(self):
        if len(self.files) < 2:
            messagebox.showwarning("Warning", "Add at least 2 Excel files")
            return

        dfs = []
        for f in self.files:
            try:
                df = pd.read_excel(f, dtype=str, header=None)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read {f}\n{e}")
                return
            dfs.append(df)

        max_rows = max(df.shape[0] for df in dfs)
        max_cols = max(df.shape[1] for df in dfs)

        wb = Workbook()
        ws = wb.active
        ws.title = "Diff"

        red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

        for r in range(max_rows):
            for c in range(max_cols):
                values = []
                for df in dfs:
                    try:
                        val = df.iat[r, c]
                        if pd.isna(val):
                            val = "—"
                    except:
                        val = "—"
                    values.append(str(val))
                cell = ws.cell(row=r + 1, column=c + 1)
                if all(v == values[0] for v in values):
                    cell.value = values[0]
                    cell.alignment = Alignment(wrap_text=True, horizontal="justify", vertical="center", shrink_to_fit=True, justifyLastLine=True)
                else:   
                    cell.value = "\n".join(f"{i + 1}: {v}" for i, v in enumerate(values))
                    cell.alignment = Alignment(wrap_text=True, horizontal="justify", vertical="center", shrink_to_fit=True, justifyLastLine=True)
                    cell.fill = red_fill

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            wb.save(save_path)
            try:
                if sys.platform.startswith("win"):
                    os.startfile(save_path)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", save_path])
            except Exception as e:
                messagebox.showinfo("Done", f"Diff saved to {save_path}\n(Couldn't open automatically: {e})")


if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = ExcelDiffTool(root)
    root.mainloop()